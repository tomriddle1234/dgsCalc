# encoding=utf8

import fixutf8
import sys,os
import logging
import re
import collections

from openpyxl import load_workbook

import argparse

from dgsConstants import *
from dgsUtil import *

codePattern = re.compile("^[A-H][0-1][0-9]-[2][0]-[0-9][0-9][0-9]-[0-9][0-9][0-9][0-9]")

parser = argparse.ArgumentParser(description='This program is to check file existence for DGS. Also generating a log file with the name of the output, end with .log ')
parser.add_argument('-i','--input', help='xlsx file path, the detail form.', required=True)
parser.add_argument('-o','--output', help='CSV file with a list of code of works', required=True)
args = vars(parser.parse_args())

xlFilename = args['input']
outputxlFilename = ""
#xlFilename = 'prepared2.xlsx'
if isinstance(args['input'],unicode):
	outputxlFilename = args['input']
else:
	outputxlFilename = args['input'].decode('gb2312').encode('utf-8')
outputCodeListFile = args['output']
logFilePath = os.path.splitext(outputCodeListFile)[0] + '.log'

print "输入Excel文件绝对路径：%s" % outputxlFilename
print "输出编号列表文件：%s" % outputCodeListFile
print "记录文件: %s" % logFilePath
### Logging
#Function : Generate a log and report file.



logging.basicConfig(filename=logFilePath, level=logging.DEBUG,format='%(asctime)s %(message)s')

logging.info('开始机检。%s' % gbk2utf8(xlFilename))

if not os.path.isfile(xlFilename):
    print "Input %s does not exist" % xlFilename
        
#if read only there will be bugs on loading cell values
wb = load_workbook(xlFilename)

sheetNameList = wb.get_sheet_names()

# This must find the 作品明细表
wsIndex = sheetNameList.index("作品明细表")
if wsIndex == None:
    outstr = "错误：xlsx文件未见 作品明细表"
    print outstr
    logging.warning(outstr)
    
ws = wb.worksheets[wsIndex]
              
def utilIsAlphabeticalOrder(wordList):
    for i in range(len(wordList) - 1):
        if wordList[i] > wordList[i+1]:
            return False
    return True

def utilGetTotalRecords(ws):
    row_count = ws.max_row
    column_count = ws.max_column
    valid_count = 0
    result = 0
    for i in range(3,row_count):     
        if  ws.cell(row=i,column=2).value is None or ws.cell(row=i,column=1).value == "汇总":
            break
        result += 1
     
#Function : Check if the work total numbers are calculated right
#openpyxl does not support calculate formula

#    for i in range(3,row_count):     
#        if  ws.cell(row=i,column=3).value.strip() == "作品数合计：":
#            given = int(ws.cell(row=i,column=4).value)
#            if given != result:
#                outstr = "机检作品数合计 %d 与表格给出数据 %d 不符，检查第 %d 排第4列。" % (result,given,i)
#                print outstr
#                logging.warning(outstr)
                     
    return result

recordsCount = utilGetTotalRecords(ws)

def utilCheckIfAllChinese(input_str):
    """
        Reuqire input is a utf8 string.
    """
    if not all((u'\u4e00' <= c <= u'\u9fff') or (c == ',') or (c == u'，') for c in input_str.decode('utf-8')):
        return False
    return True

def utilCheckIfCellEmpty(input_cell,s,i):
    if input_cell.value is None:
        outstr = "第 %d 排 %s 是空白的。" % (i, s)
        print outstr
        logging.warning(outstr)
        return False
    return True
        
     
    

def checkCategoryName(ws):
    """
    Function : Check if category name is in given list of strings
    """
    row_count = ws.max_row
    column_count = ws.max_column
    
    result = True
    for i in range(3,3+recordsCount):     
        utilCheckIfCellEmpty(ws.cell(row=i,column=2),"命题类别",i)
        # if there is a category name not in the list log this
        # Note there must be an end label as 汇总
        if not ws.cell(row=i,column=2).value.strip() in categoryList:
            outstr = "命题类别第 %s 排,第 %s 列不在类别名称列表中。" % (str(i),'2')
            print outstr
            logging.warning(outstr)
            result = False
    return result 


def checkTeamMemberNames(ws):
    """
    Function : Check if team member and teacher's names in right format
    """
    row_count = ws.max_row
    column_count = ws.max_column
    
    result = True
    for i in range(3,3+recordsCount):     
        utilCheckIfCellEmpty(ws.cell(row=i,column=7),"作者",i)
        # if name contains spaces
        if (' ' in ws.cell(row=i,column=7).value) or ('　' in ws.cell(row=i,column=7).value):
            outstr = "第 %d　排作者名称中出现空格。" % i
            print outstr
            logging.warning(outstr)
            result = False
        elif '、' in ws.cell(row=i,column=7).value:
            outstr = "第 %d　排作者名称中出现顿号。" % i
            print outstr
            logging.warning(outstr)
            result = False
        elif not utilCheckIfAllChinese(ws.cell(row=i,column=7).value):
            outstr = "第 %d　排指导教师名称中出现其它特殊符号。" % i
            print outstr
            logging.warning(outstr)
            result = False
    return result 

def checkTeacherMemberNames(ws):
    """
    Function : Check if team member and teacher's names in right format
    """
    row_count = ws.max_row
    column_count = ws.max_column
    
    result = True
    for i in range(3,3+recordsCount):     
        if utilCheckIfCellEmpty(ws.cell(row=i,column=9),"指导教师",i):
        # if name contains spaces
        
            if (u'\x20' in ws.cell(row=i,column=9).value) or (u'\u3000' in ws.cell(row=i,column=9).value):
                outstr = "第 %d　排指导教师名称中出现空格。" % i
                print outstr
                logging.warning(outstr)
                result = False
            elif  '、' in ws.cell(row=i,column=9).value:
                outstr = "第 %d　排指导教师名称中出现顿号。" % i
                print outstr
                logging.warning(outstr)
                result = False  
            elif not utilCheckIfAllChinese(ws.cell(row=i,column=9).value):
                outstr = "第 %d　排指导教师名称中出现其它特殊符号。" % i
                print outstr
                logging.warning(outstr)
                result = False
    return result 


def checkNewFormat(ws):
    """
    Function : Check if this table is the new format or the old format
    
    Check new or old by the second row
    """
    row_count = ws.max_row
    column_count = ws.max_column
	
    if not row_count or not column_count:
        outstr = "表格空白，可能未正确打开。"
        print outstr
        logging.warning(outstr)
	
    rowelements = []
    result = True
    for i in range(1,column_count):
        rowelements.append(ws.cell(row=2,column=i).value)
    
    for ele in rowelements:
        if "师2" in ele:
            result = False
            break
    
    if "系列件数" not in rowelements:
        result = False
    
    if not result:
        outstr = ">>>此表格是旧表。<<<"
        print rowelements
        print row_count
        print column_count
    else:
        outstr = "此表格是新表。"

    print outstr
    logging.warning(outstr)
    
    

def checkEncodedNumberFormat(ws):
    """    
    Function : Check if work encoded number is in right format
    Function : Check if work category is following order. 
    Function : Check if work encoded number is following order.
    """    
    
    cacheCatList = []
    cacheThemeList = []
    cacheUniCodeList = []
    cacheLastCodeList = []
    
    skipCount = 0
    
    for i in range(3,3+recordsCount):     
        
        utilCheckIfCellEmpty(ws.cell(row=i,column=4),"参赛编号",i)
        codeStr = ws.cell(row=i,column=4).value
        if not codePattern.match(codeStr):
            outstr = "参赛编号：第%d排 编号格式不正确。" % i
            print outstr
            logging.warning(outstr)
            skipCount += 1
            continue
        
        subCodeStrList = codeStr.split('-')
        
        
        
        #check first word
        #get index of category name 
        if ws.cell(row=i,column=2).value.strip() in categoryList:
            catIndex = categoryList.index(ws.cell(row=i,column=2).value.strip())
        else:
            outstr = "参赛编号：第%d排 校验编号时发现命题类别单元不在命题类别列表中。" % i   
            print outstr
            logging.warning(outstr)
        if categoryNoList[catIndex] != subCodeStrList[0][0]:
            #print subCodeStrList[0][0]
            outstr = "参赛编号：第%d排 编号中命题类别代码与命题类别不符。" % i   
            print outstr
            logging.warning(outstr)
        #get index of theme name
        if ws.cell(row=i,column=3).value.strip() in themeList:
            themeIndex = themeList.index(ws.cell(row=i,column=3).value.strip())
        else:
            outstr = "参赛编号：第%d排 校验编号时发现命题名称单元不在命题名称列表中。" % i   
            print outstr
            logging.warning(outstr)
        if themeNoList[themeIndex] != int(subCodeStrList[0][-2:]):
            outstr = "参赛编号：第%d排 编号中命题名称代码与命题名称不符。" % i   
            print outstr
            logging.warning(outstr)
            
            
        #TODO: check university code here
        
        
        #check code wether follow order
        
        #if sub string number isn't right, skip this one
        if len(subCodeStrList) != 4:
            print "This should not happen, sub string number is incorrect."
            continue
            
        
        #put sub strings in to separated list
        cacheCatList.append(subCodeStrList[0][0])
        cacheThemeList.append(int(subCodeStrList[0][-2:]))
        cacheUniCodeList.append(int(subCodeStrList[2]))
        cacheLastCodeList.append(int(subCodeStrList[3]))
    
    #all cache list length must be the same, cause records count is the same    
    assert len(cacheCatList) == len(cacheThemeList) == len(cacheUniCodeList) == len(cacheLastCodeList) == recordsCount - skipCount,"Fatal Problem on checking order or the code word. cache list lengths are not equal to record count."
    
    #start check order 
    if not utilIsAlphabeticalOrder(cacheCatList):
        outstr = ">>>命题类别参赛编号顺序：编号中命题 类别 代码顺序不对。<<<"  
        print outstr
        logging.warning(outstr)
    else:
        for ele in sorted(set(cacheCatList)): #Unique list
            subCacheThemeNoList = []
            for i in range(len(cacheCatList)):
                if cacheCatList[i] == ele: #only if category is this one
                    subCacheThemeNoList.append(cacheThemeList[i])
            if not utilIsAlphabeticalOrder(subCacheThemeNoList):
                outstr = ">>>命题类别参赛编号顺序：编号命题类别 %s 后 名称 代码顺序不对。<<<" % ele
                print outstr
                logging.warning(outstr)
            
            #category code ordered, theme code ordered, now check the work number  
            else:
            
                for item in sorted(set(subCacheThemeNoList)):
                    subCacheWorkNoList = []
                    for i in range(len(cacheCatList)):
                        if (cacheCatList[i] == ele) and (cacheThemeList[i] == item): #only if category and theme is this
                            subCacheWorkNoList.append(cacheLastCodeList[i])
                        if not utilIsAlphabeticalOrder(subCacheWorkNoList):
                            outstr = ">>>命题类别参赛编号顺序：编号命题类别 %s 名称 %s 后 作品 代码顺序不对。<<<" % (ele, item)
                            print outstr
                            logging.warning(outstr) 
            
            
        
        
        
def generateFileName(ws,filename):
    """
    Generate a list of filename can be used to search file
    """
    filenameList = []
    for i in range(3,3+recordsCount):     
        utilCheckIfCellEmpty(ws.cell(row=i,column=4),"参赛编号",i)
        utilCheckIfCellEmpty(ws.cell(row=i,column=6),"系列件数",i)
        codeStr = ws.cell(row=i,column=4).value.strip()
        itemNoStr = ws.cell(row=i,column=6).value
        filenameList.append(codeStr)
        if itemNoStr and int(itemNoStr) != 1:
            for j in range(int(itemNoStr)):
                filenameList.append(codeStr+'-'+str(j+1))
    
    
    #Check if there is duplicated names
    
    dupList = [item for item, count in collections.Counter(filenameList).items() if count > 1]
    
    if dupList:
        outstr = ">>>参赛编号：存在重复编号。列表如下：<<<"
        print outstr
        logging.warning(outstr)
        for d in dupList:
            logging.warning(d)

    
    
    
    #print filenameList
    
    writecsv(filenameList,filename)

    return
    



        
    

    
    
    
    
#Test

checkNewFormat(ws)
logging.info("")
logging.info("########################")
logging.info("")
checkCategoryName(ws)
logging.info("")
logging.info("########################")
logging.info("")
checkTeamMemberNames(ws)
logging.info("")
logging.info("########################")
logging.info("")
checkTeacherMemberNames(ws)
logging.info("")
logging.info("########################")
logging.info("")
checkEncodedNumberFormat(ws)
logging.info("")
logging.info("########################")
logging.info("")
generateFileName(ws,outputCodeListFile)

logging.info('机检完毕。%s' % gbk2utf8(xlFilename))





#Function : Label wrong records or cells with related color

#Function : Generate a statistic table and compare it with the given statistics


